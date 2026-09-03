"""MapBiomas — pressão de ocupação urbana no eixo Manso–Cuiabá.

Baixa o módulo urbano (Coleção 10) com estatísticas municipais e extrai,
para os municípios do eixo:

- área urbana 2024 (ha) — vegetada + não vegetada
- área urbana em faixa de drenagem baixa (≤3 m) — proxy de exposição
- variação 2014→2024 da área urbana (crescimento)

Saídas:
  dados/tratados/mapbiomas_pressao_eixo_cuiaba.csv
  relatorios/mapbiomas_pressao_eixo.md

Uso:
  python scripts/41_mapbiomas_eixo.py
  python executar.py 41
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any

import comum

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl é necessário: pip install openpyxl") from exc

SAIDA = comum.DADOS_TRATADOS / "mapbiomas_pressao_eixo_cuiaba.csv"
REL = comum.RELATORIOS / "mapbiomas_pressao_eixo.md"
BRUTOS = comum.DADOS_BRUTOS / "mapbiomas"
INTERESSE = comum.DADOS_TRATADOS / "cuiaba_municipios_de_interesse.json"

# Módulo urbano Coleção 10 (Google Drive — arquivo ~57 MB encapsulado)
URL_URBANO = (
    "https://drive.usercontent.google.com/download"
    "?id=1Bxa_irBxpT7gzBXj1nzpr-uQD8tergOO&export=download&authuser=0&confirm=t"
)
ZIP_NOME = "mapbiomas_col10_modulo_urbano.zip"

CAMPOS = [
    "municipio",
    "codigo_ibge",
    "bioma",
    "area_urbana_2024_ha",
    "area_urbana_2014_ha",
    "delta_urbana_10a_ha",
    "area_urbana_drenagem_ate_3m_2024_ha",
    "pct_urbana_em_drenagem_baixa",
    "fonte",
    "colecao",
    "observacao",
]


def municipios_eixo() -> dict[str, str]:
    if not INTERESSE.exists():
        return {}
    data = json.loads(INTERESSE.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    for m in data.get("municipios") or []:
        cod = str(m.get("codigo_ibge") or "").strip()
        nome = str(m.get("nome") or "").strip()
        if cod and nome:
            out[cod] = nome
    return out


def garantir_xlsx() -> Path:
    BRUTOS.mkdir(parents=True, exist_ok=True)
    zip_path = BRUTOS / ZIP_NOME
    if not zip_path.exists() or zip_path.stat().st_size < 1_000_000:
        print("  baixando módulo urbano MapBiomas Col.10…", flush=True)
        import urllib.request

        req = urllib.request.Request(
            URL_URBANO,
            headers={"User-Agent": "VIGIBARRAGENS-MT/1.0 (MapBiomas eixo)"},
        )
        with urllib.request.urlopen(req, timeout=180) as resp, zip_path.open("wb") as out:
            while True:
                chunk = resp.read(1024 * 256)
                if not chunk:
                    break
                out.write(chunk)
        print(f"  gravado {zip_path.relative_to(comum.RAIZ)}")

    # Extrai xlsx interno
    with zipfile.ZipFile(zip_path) as zf:
        membros = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not membros:
            raise SystemExit("zip MapBiomas sem xlsx")
        membro = membros[0]
        destino = BRUTOS / Path(membro).name
        if not destino.exists() or destino.stat().st_size < 100_000:
            with zf.open(membro) as src, destino.open("wb") as dst:
                dst.write(src.read())
            print(f"  extraído {destino.name}")
        return destino


def _f(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def agregar(xlsx: Path, munis: dict[str, str]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    cods = set(munis)
    # UrbanVegetation: soma classes urbanas
    urbana: dict[str, dict[str, Any]] = {}
    ws = wb["UrbanVegetation"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        cod = str(int(row[0])) if str(row[0]).isdigit() or isinstance(row[0], (int, float)) else str(row[0])
        cod = "".join(ch for ch in cod if ch.isdigit())
        if len(cod) == 6:
            # geocódigo sem dígito verificador — tenta casar prefixo
            hits = [c for c in cods if c.startswith(cod)]
            if not hits:
                continue
            cod = hits[0]
        if cod not in cods:
            continue
        class_nm = str(row[5] or "")
        if class_nm not in {"Vegetated Urban Area", "Non-Vegetated Urban Area"}:
            continue
        try:
            v2014 = _f(row[6 + (2014 - 1985)])
            v2024 = _f(row[6 + (2024 - 1985)])
        except IndexError:
            continue
        slot = urbana.setdefault(
            cod,
            {
                "municipio": munis[cod],
                "codigo_ibge": cod,
                "bioma": str(row[4] or ""),
                "area_urbana_2014_ha": 0.0,
                "area_urbana_2024_ha": 0.0,
                "area_urbana_drenagem_ate_3m_2024_ha": 0.0,
            },
        )
        slot["area_urbana_2014_ha"] += v2014
        slot["area_urbana_2024_ha"] += v2024

    # HeightDrainage: área urbana em ≤3 m (proxy de ocupação em terreno baixo)
    if "HeightDrainage" in wb.sheetnames:
        ws2 = wb["HeightDrainage"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            cod = "".join(ch for ch in str(row[0]) if ch.isdigit())
            if len(cod) == 6:
                hits = [c for c in cods if c.startswith(cod)]
                if not hits:
                    continue
                cod = hits[0]
            if cod not in urbana and cod not in cods:
                continue
            if cod not in urbana:
                urbana[cod] = {
                    "municipio": munis[cod],
                    "codigo_ibge": cod,
                    "bioma": str(row[4] or ""),
                    "area_urbana_2014_ha": 0.0,
                    "area_urbana_2024_ha": 0.0,
                    "area_urbana_drenagem_ate_3m_2024_ha": 0.0,
                }
            class_nm = str(row[5] or "")
            # classes: "Até 3m", "Entre 3 e 6m", "Maior que 6m"
            if not (class_nm.startswith("Até") or class_nm.startswith("Ate")):
                continue
            if "3" not in class_nm:
                continue
            try:
                v2024 = _f(row[6 + (2024 - 1985)])
            except IndexError:
                continue
            urbana[cod]["area_urbana_drenagem_ate_3m_2024_ha"] += v2024

    out: list[dict[str, Any]] = []
    for cod, nome in munis.items():
        slot = urbana.get(cod) or {
            "municipio": nome,
            "codigo_ibge": cod,
            "bioma": "",
            "area_urbana_2014_ha": 0.0,
            "area_urbana_2024_ha": 0.0,
            "area_urbana_drenagem_ate_3m_2024_ha": 0.0,
        }
        u24 = float(slot["area_urbana_2024_ha"])
        u14 = float(slot["area_urbana_2014_ha"])
        dren = float(slot["area_urbana_drenagem_ate_3m_2024_ha"])
        pct = round(100.0 * dren / u24, 1) if u24 > 0 else 0.0
        out.append(
            {
                "municipio": slot["municipio"],
                "codigo_ibge": cod,
                "bioma": slot.get("bioma") or "",
                "area_urbana_2024_ha": round(u24, 1),
                "area_urbana_2014_ha": round(u14, 1),
                "delta_urbana_10a_ha": round(u24 - u14, 1),
                "area_urbana_drenagem_ate_3m_2024_ha": round(dren, 1),
                "pct_urbana_em_drenagem_baixa": pct,
                "fonte": "MapBiomas",
                "colecao": "10_modulo_urbano",
                "observacao": "Pressão de ocupação municipal — não é mancha na faixa HAND",
            }
        )
    out.sort(key=lambda r: -float(r["area_urbana_2024_ha"]))
    return out


def main() -> None:
    comum.preparar_diretorios()
    munis = municipios_eixo()
    if not munis:
        raise SystemExit("municípios do eixo ausentes — rode etapa 12")
    print("MapBiomas pressão urbana no eixo…", flush=True)
    xlsx = garantir_xlsx()
    regs = agregar(xlsx, munis)
    comum.salvar_csv(SAIDA, regs, CAMPOS)
    tot_u = sum(float(r["area_urbana_2024_ha"]) for r in regs)
    tot_d = sum(float(r["area_urbana_drenagem_ate_3m_2024_ha"]) for r in regs)
    REL.write_text(
        "\n".join(
            [
                "# MapBiomas — pressão de ocupação (eixo Manso–Cuiabá)",
                "",
                f"- Municípios: **{len(regs)}**",
                f"- Área urbana 2024 (soma): **{tot_u:,.0f} ha**".replace(",", "."),
                f"- Urbana em drenagem ≤3 m: **{tot_d:,.0f} ha**".replace(",", "."),
                f"- Arquivo: `{SAIDA.relative_to(comum.RAIZ)}`",
                f"- Fonte: módulo urbano Coleção 10 (`{URL_URBANO[:60]}…`)",
                "",
                "Uso: contexto de exposição na faixa de atenção; não entra no IDAP numérico.",
                "A parcela em drenagem baixa (≤3 m) é proxy de ocupação em terreno sensível a cheia.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(f"  gravado {REL.relative_to(comum.RAIZ)}")
    print(f"  urbana={tot_u:.0f} ha · drenagem≤3m={tot_d:.0f} ha")


if __name__ == "__main__":
    main()
